import streamlit as st
import pandas as pd
import numpy as np
import tempfile
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


st.title("📘 Unificador de Notas – 1º, 2º e 3º Bimestres (Notas Vermelhas < 5)")


# --------------------------------------------------------------
#  FUNÇÃO PARA DETECTAR SE O TEXTO É UM NOME DE ALUNO REAL
# --------------------------------------------------------------

def eh_aluno(nome):
    if pd.isna(nome):
        return False

    partes = str(nome).split()

    if len(partes) < 2:
        return False

    if not all(p.isalpha() for p in partes):
        return False

    if len(partes[0]) <= 2:  # remove EP, ES, ET, AC...
        return False

    return True


# --------------------------------------------------------------
#  LIMPEZA DAS PLANILHAS
# --------------------------------------------------------------

def limpar_planilha(file):
    df_raw = pd.read_excel(file, header=None)

    try:
        linha_cab = df_raw[df_raw.iloc[:, 0] == "ALUNO"].index[0]
    except:
        st.error("❌ A planilha enviada não contém a coluna 'ALUNO'.")
        st.stop()

    df = pd.read_excel(file, header=linha_cab)

    df = df[df["ALUNO"].apply(eh_aluno)]

    df = df.loc[:, ~df.columns.str.contains("Unnamed")]
    df = df.drop(columns=["SITUAÇÃO", "TOTAL"], errors="ignore")

    materias_proibidas = ["arte", "esporte", "música", "artes", "big a", "inovação", "inovacao"]

    def coluna_proibida(col):
        texto = col.lower()
        return any(p in texto for p in materias_proibidas)

    df = df[[c for c in df.columns if not coluna_proibida(c)]]

    def extrair_nota(valor):
        if pd.isna(valor):
            return np.nan
        nums = re.findall(r"\d+", str(valor))
        if not nums:
            return np.nan
        n = int(nums[0])
        return n if 0 <= n <= 10 else np.nan

    colunas_validas = ["ALUNO"]
    renomear = {}

    for col in df.columns:
        if col == "ALUNO":
            continue

        df[col] = df[col].apply(extrair_nota)

        if df[col].notna().sum() > 0:
            colunas_validas.append(col)
        else:
            continue

        # Tenta pegar o nome da materia antes do numero
        materia = re.split(r"\d+", col)[0].strip().lower() 
        
        # Limpa possíveis separadores no nome da matéria
        materia = materia.replace('-', '').replace('.', '').replace('/', '').strip()


        renomear[col] = materia

    df = df[colunas_validas]
    df = df.rename(columns=renomear)

    return df


# --------------------------------------------------------------
#  FORMATAÇÃO DO CABEÇALHO EM 2 LINHAS (COM CÉLULAS MESCLADAS)
# --------------------------------------------------------------

def formatar_cabecalho_simples(path, df_final):
    wb = load_workbook(path)
    ws = wb.active

    # O df_final foi escrito com startrow=0, entao a L1 tem o cabeçalho do DF ("ALUNO", "matéria_B1", etc.)
    # Removemos a linha 1 (cabeçalho padrão do DF).
    ws.delete_rows(1)
    
    # Inserimos as 2 linhas para o novo cabeçalho (Matéria e Bimestre)
    ws.insert_rows(1)
    ws.insert_rows(2)

    ws["A1"] = "ALUNO"
    ws["A2"] = ""

    # Dicionário para rastrear as colunas por matéria
    materias_e_colunas = {}
    
    # 1. Popula o dicionário de rastreamento de colunas e escreve o Bimestre (Linha 2)
    col_excel = 2
    for col in df_final.columns:
        if col == "ALUNO":
            continue

        partes = col.split("_")
        materia = partes[0]
        bi = partes[1]

        if materia not in materias_e_colunas:
            materias_e_colunas[materia] = []
        
        # Adiciona o número da coluna (2, 3, 4, etc.)
        materias_e_colunas[materia].append(col_excel) 

        # Linha 2: Bimestre formatado (Ex: B1 -> 1º Bi)
        bimestre_formatado = bi.replace("B", "") + "º Bi" 
        ws.cell(row=2, column=col_excel, value=bimestre_formatado)
        
        col_excel += 1

    # 2. Mescla as células e escreve a Matéria (Linha 1)
    for materia, colunas in materias_e_colunas.items():
        primeira_col = colunas[0]
        ultima_col = colunas[-1]
        
        # Converte o número da coluna para a letra (Ex: 2 -> B)
        col_inicio_letra = get_column_letter(primeira_col)
        col_fim_letra = get_column_letter(ultima_col)

        # Mescla as células da Linha 1 (Ex: B1:D1)
        ws.merge_cells(f'{col_inicio_letra}1:{col_fim_letra}1')
        
        # Escreve o nome da Matéria na primeira célula mesclada
        ws.cell(row=1, column=primeira_col, value=materia.capitalize())

        # Centraliza o texto na célula mesclada
        ws.cell(row=1, column=primeira_col).alignment = Alignment(horizontal="center", vertical="center")


    # Formata a célula ALUNO
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    wb.save(path)


# --------------------------------------------------------------
#  UPLOAD DOS 3 BIMESTRES
# --------------------------------------------------------------

file_b1 = st.file_uploader("📤 Envie o Excel do 1º Bimestre", type=["xlsx"])
file_b2 = st.file_uploader("📤 Envie o Excel do 2º Bimestre", type=["xlsx"])
file_b3 = st.file_uploader("📤 Envie o Excel do 3º Bimestre", type=["xlsx"])

if file_b1 and file_b2 and file_b3:

    st.success("Arquivos carregados! Processando...")

    df1 = limpar_planilha(file_b1)
    df2 = limpar_planilha(file_b2)
    df3 = limpar_planilha(file_b3)

    # Captura a ordem dos alunos do primeiro bimestre para manter a ordenação
    ordem_b1 = df1["ALUNO"].tolist()

    # Renomeia as colunas de notas para identificar o bimestre
    df1 = df1.rename(columns={c: f"{c}_B1" for c in df1.columns if c != "ALUNO"})
    df2 = df2.rename(columns={c: f"{c}_B2" for c in df2.columns if c != "ALUNO"})
    df3 = df3.rename(columns={c: f"{c}_B3" for c in df3.columns if c != "ALUNO"})

    # Unifica os DataFrames
    df_final = df1.merge(df2, on="ALUNO", how="outer")
    df_final = df_final.merge(df3, on="ALUNO", how="outer")

    # Preenche NaN com traço
    df_final = df_final.fillna("–")

    # Reordena pela lista do 1º bimestre
    df_final["ordem"] = df_final["ALUNO"].apply(
        lambda nome: ordem_b1.index(nome) if nome in ordem_b1 else 999
    )
    df_final = df_final.sort_values("ordem").drop(columns=["ordem"])

    st.subheader("📄 Planilha Final (antes da coloração)")
    st.dataframe(df_final)

    # Salva o DataFrame em um arquivo temporário, começando na Linha 1
    temp_out = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    df_final.to_excel(temp_out.name, index=False, startrow=0) 

    # Aplica o cabeçalho de 2 linhas e as correções de estrutura
    formatar_cabecalho_simples(temp_out.name, df_final)

    def colorir_notas(path):
        wb = load_workbook(path)
        ws = wb.active
        red = Font(color="FF0000", bold=True)

        # Começa a colorir a partir da Linha 3 (onde os dados do aluno começam agora)
        for col in range(2, ws.max_column + 1):
            for row in range(3, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                try:
                    if isinstance(val, (int, float)) and val < 5:
                        ws.cell(row=row, column=col).font = red
                except:
                    pass

        wb.save(path)

    # Colore as notas vermelhas
    colorir_notas(temp_out.name)

    # Botão de download
    with open(temp_out.name, "rb") as f:
        st.download_button(
            "⬇️ Baixar Planilha Final (Formatada + Notas Vermelhas)",
            f.read(),
            file_name="notas_unificadas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
